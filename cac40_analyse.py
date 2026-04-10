"""
CAC 40 — Analyse Financière Automatisée
Auteur : Alain Faya Somadouno | Master Finance, Y Schools SCBS Troyes
Description : Récupère les données financières en temps réel via Yahoo Finance,
              calcule les ratios clés et exporte un fichier Excel formaté.
Librairies requises : pip install yfinance pandas openpyxl
"""

import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime
import warnings
warnings.filterwarnings("ignore")

# ── 1. TICKERS CAC 40 (Yahoo Finance) ─────────────────────────────────────────

TICKERS = {
    "LVMH":             "MC.PA",
    "TotalEnergies":    "TTE.PA",
    "Hermès":           "RMS.PA",
    "Airbus":           "AIR.PA",
    "Schneider Elec.":  "SU.PA",
    "L'Oréal":          "OR.PA",
    "BNP Paribas":      "BNP.PA",
    "AXA":              "CS.PA",
    "Sanofi":           "SAN.PA",
    "Kering":           "KER.PA",
    "Société Générale": "GLE.PA",
    "Renault":          "RNO.PA",
    "Stellantis":       "STLAM.AS",
    "Capgemini":        "CAP.PA",
    "Michelin":         "ML.PA",
    "Danone":           "BN.PA",
    "Engie":            "ENGI.PA",
    "Saint-Gobain":     "SGO.PA",
    "Orange":           "ORA.PA",
    "Pernod Ricard":    "RI.PA",
}

SECTEURS = {
    "LVMH": "Luxe", "Hermès": "Luxe", "Kering": "Luxe",
    "TotalEnergies": "Énergie", "Engie": "Énergie",
    "Airbus": "Industrie", "Schneider Elec.": "Industrie",
    "Michelin": "Industrie", "Saint-Gobain": "Industrie",
    "L'Oréal": "Cosmétique", "Sanofi": "Pharma",
    "BNP Paribas": "Banque", "Société Générale": "Banque",
    "AXA": "Assurance", "Renault": "Automobile", "Stellantis": "Automobile",
    "Capgemini": "Tech / Conseil", "Danone": "Agro-alimentaire",
    "Orange": "Télécom", "Pernod Ricard": "Luxe / Spiritueux",
}

# ── 2. RÉCUPÉRATION DES DONNÉES ────────────────────────────────────────────────

def get_financial_data(name, ticker_symbol):
    """Récupère les données financières d'une entreprise via Yahoo Finance."""
    try:
        ticker = yf.Ticker(ticker_symbol)
        info   = ticker.info

        # Données brutes (en milliards €)
        def to_mrd(val):
            if val and val != "N/A":
                return round(val / 1e9, 2)
            return None

        ca          = to_mrd(info.get("totalRevenue"))
        ebitda      = to_mrd(info.get("ebitda"))
        rn          = to_mrd(info.get("netIncomeToCommon"))
        cap         = to_mrd(info.get("marketCap"))
        dettes      = to_mrd(info.get("totalDebt"))
        effectif    = info.get("fullTimeEmployees")
        cours       = info.get("currentPrice")
        beta        = info.get("beta")

        # Ratios calculés
        marge_nette  = round(rn  / ca    * 100, 1) if ca   and rn    else None
        marge_ebitda = round(ebitda / ca  * 100, 1) if ca   and ebitda else None
        levier       = round(dettes / ebitda,   1) if ebitda and dettes else None
        pe           = round(cap / rn,          1) if rn    and cap    else None
        ev_ebitda    = round((cap + dettes) / ebitda, 1) if cap and dettes and ebitda else None

        return {
            "Entreprise":       name,
            "Secteur":          SECTEURS.get(name, "Autre"),
            "Ticker":           ticker_symbol,
            "CA (Mrd €)":       ca,
            "EBITDA (Mrd €)":   ebitda,
            "Résultat Net (Mrd €)": rn,
            "Capitalisation (Mrd €)": cap,
            "Dettes (Mrd €)":   dettes,
            "Effectif":         effectif,
            "Cours (€)":        cours,
            "Bêta":             beta,
            "Marge nette (%)":  marge_nette,
            "Marge EBITDA (%)": marge_ebitda,
            "Levier (x)":       levier,
            "P/E (x)":          pe,
            "EV/EBITDA (x)":    ev_ebitda,
        }

    except Exception as e:
        print(f"  ⚠️  Erreur pour {name} ({ticker_symbol}) : {e}")
        return None


def fetch_all():
    """Récupère les données pour toutes les entreprises."""
    print("\n📊 Récupération des données CAC 40 en cours...\n")
    results = []
    for name, ticker in TICKERS.items():
        print(f"  → {name} ({ticker})")
        data = get_financial_data(name, ticker)
        if data:
            results.append(data)
    print(f"\n✅ {len(results)}/{len(TICKERS)} entreprises récupérées\n")
    return pd.DataFrame(results)

# ── 3. ANALYSE ET STATISTIQUES ─────────────────────────────────────────────────

def analyse_par_secteur(df):
    """Calcule les moyennes par secteur."""
    cols_num = ["CA (Mrd €)", "Capitalisation (Mrd €)",
                "Marge nette (%)", "Marge EBITDA (%)", "Levier (x)", "EV/EBITDA (x)"]
    synthese = df.groupby("Secteur")[cols_num].mean().round(2).reset_index()
    synthese.insert(1, "Nb entreprises", df.groupby("Secteur").size().values)
    return synthese


def top_entreprises(df, ratio, n=5, ascending=False):
    """Retourne le top N entreprises selon un ratio."""
    return (df[["Entreprise", "Secteur", ratio]]
            .dropna()
            .sort_values(ratio, ascending=ascending)
            .head(n)
            .reset_index(drop=True))

# ── 4. EXPORT EXCEL ───────────────────────────────────────────────────────────

BLUE_DARK  = "1F3864"
BLUE_LIGHT = "D6E4F7"
WHITE      = "FFFFFF"
GRAY       = "F5F5F5"

thin   = Side(style="thin",   color="BFBFBF")
medium = Side(style="medium", color="1F3864")
b_thin = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
b_med  = Border(left=medium, right=medium, top=medium, bottom=medium)


def style_header(cell, bg=BLUE_DARK, fg=WHITE):
    cell.font      = Font(bold=True, color=fg,  name="Arial", size=9)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = b_med


def style_data(cell, bg=WHITE, bold=False, fmt=None, halign="center"):
    cell.font      = Font(name="Arial", size=9, bold=bold)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=halign, vertical="center")
    cell.border    = b_thin
    if fmt:
        cell.number_format = fmt


def export_excel(df, synthese, filename="CAC40_Analyse_Somadouno.xlsx"):
    wb = Workbook()

    # ── Onglet 1 : Données complètes ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Données temps réel"

    date_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws1.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
    ws1["A1"] = f"ANALYSE CAC 40 — Données temps réel | Extrait le {date_str} | Alain Faya Somadouno"
    ws1["A1"].font  = Font(bold=True, color=WHITE, name="Arial", size=11)
    ws1["A1"].fill  = PatternFill("solid", start_color=BLUE_DARK)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 26

    headers = list(df.columns)
    for col, h in enumerate(headers, 1):
        style_header(ws1.cell(row=2, column=col, value=h))
    ws1.row_dimensions[2].height = 32

    pct_cols   = {"Marge nette (%)": "0.0\"%\"", "Marge EBITDA (%)": "0.0\"%\""}
    multi_cols = {"Levier (x)": "0.0\"x\"", "P/E (x)": "0.0\"x\"", "EV/EBITDA (x)": "0.0\"x\""}

    for r, row in enumerate(df.itertuples(index=False), 3):
        bg = GRAY if r % 2 == 0 else WHITE
        ws1.row_dimensions[r].height = 17
        for col, (h, val) in enumerate(zip(headers, row), 1):
            fmt  = pct_cols.get(h) or multi_cols.get(h)
            bold = h == "Entreprise"
            ha   = "left" if h in ("Entreprise", "Secteur") else "center"
            c    = ws1.cell(row=r, column=col, value=val)
            style_data(c, bg=bg, bold=bold, fmt=fmt, halign=ha)

    # Mise en forme conditionnelle sur marges
    n_rows = len(df) + 2
    for col_letter in ["L", "M"]:  # Marge nette, Marge EBITDA
        ws1.conditional_formatting.add(
            f"{col_letter}3:{col_letter}{n_rows}",
            ColorScaleRule(start_type="min", start_color="FCE4EC",
                           mid_type="percentile", mid_value=50, mid_color="FFFFFF",
                           end_type="max", end_color="E2EFDA")
        )

    # Largeurs colonnes
    widths = [18,16,8,9,9,12,13,9,9,8,5,11,13,9,7,10]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "D3"

    # ── Onglet 2 : Synthèse secteurs ──────────────────────────────────────────
    ws2 = wb.create_sheet("Synthèse Secteurs")
    ws2.merge_cells(f"A1:{get_column_letter(len(synthese.columns))}1")
    ws2["A1"] = "SYNTHÈSE PAR SECTEUR — Moyennes des ratios clés"
    ws2["A1"].font  = Font(bold=True, color=WHITE, name="Arial", size=11)
    ws2["A1"].fill  = PatternFill("solid", start_color=BLUE_DARK)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    for col, h in enumerate(synthese.columns, 1):
        style_header(ws2.cell(row=2, column=col, value=h))
    ws2.row_dimensions[2].height = 28

    for r, row in enumerate(synthese.itertuples(index=False), 3):
        bg = GRAY if r % 2 == 0 else WHITE
        ws2.row_dimensions[r].height = 18
        for col, (h, val) in enumerate(zip(synthese.columns, row), 1):
            fmt = "0.0\"%\"" if "%" in h else ("0.0\"x\"" if "(x)" in h else None)
            c   = ws2.cell(row=r, column=col, value=val)
            style_data(c, bg=bg, bold=(col==1), fmt=fmt,
                       halign="left" if col==1 else "center")

    for i, w in enumerate([18,13,13,15,14,15,10,12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Onglet 3 : Top classements ────────────────────────────────────────────
    ws3 = wb.create_sheet("Top Classements")
    ws3.merge_cells("A1:D1")
    ws3["A1"] = "TOP 5 PAR RATIO — Identifier les leaders par indicateur"
    ws3["A1"].font  = Font(bold=True, color=WHITE, name="Arial", size=11)
    ws3["A1"].fill  = PatternFill("solid", start_color=BLUE_DARK)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 26

    classements = [
        ("Meilleures marges nettes",     "Marge nette (%)",  False, "0.0\"%\""),
        ("Meilleures marges EBITDA",     "Marge EBITDA (%)", False, "0.0\"%\""),
        ("Levier le plus faible (solide)","Levier (x)",      True,  "0.0\"x\""),
        ("Meilleur EV/EBITDA (attractif)","EV/EBITDA (x)",   True,  "0.0\"x\""),
    ]

    current_row = 3
    for title, ratio, asc, fmt in classements:
        ws3.merge_cells(f"A{current_row}:D{current_row}")
        title_cell = ws3.cell(row=current_row, column=1, value=title)
        title_cell.font  = Font(bold=True, color=WHITE, name="Arial", size=9)
        title_cell.fill  = PatternFill("solid", start_color="2F5496")
        title_cell.alignment = Alignment(horizontal="left", vertical="center",
                                         indent=1)
        ws3.row_dimensions[current_row].height = 20
        current_row += 1

        top = top_entreprises(df, ratio, ascending=asc)
        for col_h, h in enumerate(["#", "Entreprise", "Secteur", ratio], 1):
            style_header(ws3.cell(row=current_row, column=col_h, value=h),
                         bg="BDD7EE", fg="1F3864")
        ws3.row_dimensions[current_row].height = 20
        current_row += 1

        for rank, row in enumerate(top.itertuples(index=False), 1):
            bg = GRAY if rank % 2 == 0 else WHITE
            ws3.row_dimensions[current_row].height = 17
            ws3.cell(row=current_row, column=1, value=rank).font = Font(
                bold=True, name="Arial", size=9, color="2F5496")
            ws3.cell(row=current_row, column=1).alignment = Alignment(horizontal="center")
            ws3.cell(row=current_row, column=1).fill = PatternFill("solid", start_color=bg)
            for col, val in enumerate(row, 2):
                c = ws3.cell(row=current_row, column=col, value=val)
                style_data(c, bg=bg, fmt=fmt if col==4 else None,
                           halign="left" if col <= 3 else "center")
            current_row += 1
        current_row += 1

    for i, w in enumerate([5, 20, 18, 16], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save(filename)
    print(f"✅ Fichier exporté : {filename}")
    return filename

# ── 5. MAIN ───────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    # Récupération des données
    df = fetch_all()

    if df.empty:
        print("❌ Aucune donnée récupérée. Vérifie ta connexion internet.")
    else:
        # Affichage synthétique dans le terminal
        print("=" * 65)
        print("RÉSUMÉ — CAC 40 | Top marges nettes")
        print("=" * 65)
        top5 = top_entreprises(df, "Marge nette (%)")
        print(top5.to_string(index=False))

        print("\n" + "=" * 65)
        print("RÉSUMÉ — Levier le plus faible (entreprises les plus solides)")
        print("=" * 65)
        top5_lev = top_entreprises(df, "Levier (x)", ascending=True)
        print(top5_lev.to_string(index=False))

        # Synthèse par secteur
        synthese = analyse_par_secteur(df)

        # Export Excel
        print("\n📁 Export Excel en cours...")
        filename = export_excel(df, synthese)

        print(f"\n🚀 Script terminé avec succès !")
        print(f"   → Ouvre '{filename}' dans Power BI pour construire le dashboard")
        print(f"   → Relance ce script n'importe quand pour mettre à jour les données")
